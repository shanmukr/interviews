#!/usr/bin/python3
#import openpyxl
#from openpyxl import load_workbook
from xlrd import open_workbook
import re
import openpyxl

#xfile = openpyxl.load_workbook('sample.xlsx')
#sheet = xfile.get_sheet_by_name('Sheet1')
#sheet['C1'] = 'hello world'
#xfile.save("sample.xlsx")

def files_read(file_path, file_name):
  full_path = file_path+'parm\\'+file_name
  print( full_path)
  lines = open(full_path, 'r')
  version = []
  for i in lines.readlines():
    match_data = re.match("ver=\"(\w+)\"", i)
    if match_data:
      version.append(match_data.group(1))
  lines.closed
  return version


wb = open_workbook('sample.xlsx')
for s in wb.sheets():
  version = files_read( s.cell(0,0).value, s.cell(0,1).value )
  print (version)
