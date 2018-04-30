#import xlsxwriter
import openpyxl

#wb=openpyxl.load_workbook('sample.xlsx')
#test = wb.get_sheet_names()
#sheet=wb.get_sheet_by_name('Sheet1')
#print (test, sheet.title)
#print (sheet.title('A1'))

book = openpyxl.load_workbook('sample.xlsx')

sheet = book.active

a = sheet['O3']
#a2 = sheet['A3']
#t = sheet['A2']
#a3 = sheet.cell(row=1, column=1)

print (a.value)
