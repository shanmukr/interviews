import openpyxl

xfile = openpyxl.load_workbook('apps.xlsx')
sheet = xfile.get_sheet_by_name('Comprehensive apps info')
#sheets = xfile.get_sheet_names()

def read_data(path):
  print(path)

for i in range(3,224):
  if (sheet["X"+str(i)].value):
    read_data(sheet["X"+str(i)].value+"/parm/"+sheet["F"+str(i)].value)
  else:
    print (sheet["D"+str(i)].value, "path was not present.")


"""
def read_data(path):
  print(path)

for i in range(3,sheet.max_row+1):
  if (sheet["X"+str(i)].value):
    read_data(sheet["X"+str(i)].value)
  else:
    print (sheet["D"+str(i)].value, "path was not present.")
  
#print(sheet.max_row)
"""
"""
sheet['C1'] = 'hello world'
xfile.save("sample.xlsx")
"""
