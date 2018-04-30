from multiprocessing import Lock, Queue, Process
from re import search
from time import sleep
from os import listdir
import openpyxl

file_path = "dir_path"
no_processes = 20

def pro(q, lo):
 while lo:
  time.sleep(0.5)
 lo.acquire()
 try:
  book = q.get(timeout=20)
 except TimeoutError:
  return
 lo.release()
 sheet_names = book.get_sheet_names()
 for i in sheet_names:
  sheet = book.get_sheet_by_name(i)
  for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
   for cell in row:
    match = search('NULL|\d{2}-\d{2}\-\d{4}', cell.value, re.I) # check "Null" and date fromat
    if match:
     print(row)
     break

queue = Queue()
lock = Lock()
files = listdir(file_path)

for file in files:
 try:
  fd = openpyxl.load_workbook(file_path+file)
 except IOError:
  print("Cannot open file ", file)
  continue
 q.put(fd)

for i in range(no_processes):
 p = Process(target = pro, args = (queue, lock))
 p.start()


